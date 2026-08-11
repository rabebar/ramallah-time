from __future__ import annotations

import io
import os
import re
import sqlite3
from contextlib import contextmanager
from functools import wraps
from datetime import date, datetime
from pathlib import Path

from flask import Flask, flash, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash, generate_password_hash


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DB_PATH = (
    Path("/opt/render/project/src/static/uploads/.flex_private/flex.db")
    if os.getenv("RENDER")
    else BASE_DIR / "flex.db"
)
DB_PATH = Path(os.getenv("FLEX_DB_PATH", DEFAULT_DB_PATH))
DB_PATH.parent.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.getenv("FLEX_SECRET_KEY") or os.getenv("SECRET_KEY") or "local-flex-prototype-key"
app.config.update(
    SESSION_COOKIE_NAME="flex_session",
    SESSION_COOKIE_PATH="/flex" if os.getenv("RENDER") else "/",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=bool(os.getenv("RENDER")),
)

SERVICE_CATALOG = [
    ("ملابس", "قميص", "قطعة", 8),
    ("ملابس", "تيشيرت", "قطعة", 7),
    ("ملابس", "قميص بولو", "قطعة", 8),
    ("ملابس", "بلوزة", "قطعة", 9),
    ("ملابس", "كنزة / سويتر", "قطعة", 12),
    ("ملابس", "هودي", "قطعة", 14),
    ("ملابس", "فيست / سترة", "قطعة", 10),
    ("ملابس", "بنطال قماش", "قطعة", 10),
    ("ملابس", "جينز", "قطعة", 10),
    ("ملابس", "شورت", "قطعة", 8),
    ("ملابس", "تنورة", "قطعة", 12),
    ("ملابس", "فستان عادي", "قطعة", 25),
    ("ملابس", "فستان سهرة", "قطعة", 40),
    ("ملابس", "بدلة قطعتان", "قطعة", 30),
    ("ملابس", "بدلة ثلاث قطع", "قطعة", 40),
    ("ملابس", "جاكيت / بليزر", "قطعة", 20),
    ("ملابس", "معطف", "قطعة", 25),
    ("ملابس", "عباءة / جلابية", "قطعة", 20),
    ("ملابس", "ثوب / دشداشة", "قطعة", 15),
    ("ملابس", "وشاح / شال", "قطعة", 7),
    ("ملابس", "ربطة عنق", "قطعة", 6),
    ("مفروشات", "شرشف مفرد", "قطعة", 10),
    ("مفروشات", "شرشف مزدوج", "قطعة", 15),
    ("مفروشات", "غطاء لحاف", "قطعة", 18),
    ("مفروشات", "لحاف مفرد", "قطعة", 25),
    ("مفروشات", "لحاف مزدوج", "قطعة", 35),
    ("مفروشات", "بطانية مفرد", "قطعة", 20),
    ("مفروشات", "بطانية مزدوج", "قطعة", 30),
    ("مفروشات", "كيس وسادة", "قطعة", 5),
    ("مفروشات", "منشفة", "قطعة", 6),
    ("مفروشات", "مفرش طاولة", "قطعة", 15),
    ("سجاد وستائر", "سجاد", "متر مربع", 12),
    ("سجاد وستائر", "موكيت", "متر مربع", 10),
    ("سجاد وستائر", "ستائر خفيفة", "متر", 10),
    ("سجاد وستائر", "ستائر ثقيلة", "متر", 15),
    ("أحذية وحقائب", "حذاء رياضي", "قطعة", 20),
    ("أحذية وحقائب", "حذاء جلدي", "قطعة", 25),
    ("أحذية وحقائب", "حقيبة قماش", "قطعة", 20),
    ("أحذية وحقائب", "حقيبة جلد", "قطعة", 30),
    ("خدمات", "كوي فقط", "قطعة", 5),
    ("خدمات", "غسيل بالكيلو", "كيلو", 10),
]

GARMENT_TREATMENTS = (
    ("غسيل وكوي", 1.00),
    ("غسيل فقط", 0.75),
    ("كوي فقط", 0.55),
    ("تنظيف جاف", 1.25),
)


def service_catalog_rows():
    rows = []
    for category, name, unit, price in SERVICE_CATALOG:
        if category == "ملابس":
            for treatment, factor in GARMENT_TREATMENTS:
                calculated_price = round(max(price * factor, 1) * 2) / 2
                rows.append((category, name, treatment, unit, calculated_price))
        else:
            rows.append((category, name, "حسب الوصف", unit, price))
    return rows


def requested_due_date():
    """Read a delivery date in the explicit day/month/year order used by FLEX."""
    if not any(request.form.get(name) for name in ("due_day", "due_month", "due_year", "due_time")):
        return (request.form.get("due_date") or "").strip() or None
    day = int(request.form.get("due_day") or 0)
    month = int(request.form.get("due_month") or 0)
    year = int(request.form.get("due_year") or 0)
    time_value = (request.form.get("due_time") or "00:00").strip()
    hour, minute = (int(part) for part in time_value.split(":", 1))
    value = datetime(year, month, day, hour, minute)
    return value.strftime("%Y-%m-%dT%H:%M")


@contextmanager
def db():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    try:
        yield connection
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def delete_business_account(connection, business_id):
    """Permanently remove one FLEX tenant and all of its operational data."""
    order_ids = [
        row[0] for row in connection.execute(
            "SELECT id FROM orders WHERE business_id=?", (business_id,)
        ).fetchall()
    ]
    if order_ids:
        placeholders = ",".join("?" for _ in order_ids)
        connection.execute(f"DELETE FROM order_items WHERE order_id IN ({placeholders})", order_ids)
        connection.execute(f"DELETE FROM payments WHERE order_id IN ({placeholders})", order_ids)
    for table in ("orders", "customers", "services", "expenses", "audit_log", "cash_days", "users"):
        connection.execute(f"DELETE FROM {table} WHERE business_id=?", (business_id,))
    return connection.execute("DELETE FROM businesses WHERE id=?", (business_id,)).rowcount


def init_db():
    schema = """
    CREATE TABLE IF NOT EXISTS settings (
        key TEXT PRIMARY KEY,
        value TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS businesses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        phone TEXT,
        address TEXT,
        invoice_note TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        business_id INTEGER NOT NULL,
        full_name TEXT NOT NULL,
        phone TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'owner',
        active INTEGER NOT NULL DEFAULT 1,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(business_id) REFERENCES businesses(id)
    );
    CREATE TABLE IF NOT EXISTS services (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        category TEXT NOT NULL,
        name TEXT NOT NULL,
        treatment TEXT NOT NULL DEFAULT 'حسب الوصف',
        unit TEXT NOT NULL DEFAULT 'قطعة',
        price REAL NOT NULL DEFAULT 0,
        active INTEGER NOT NULL DEFAULT 1,
        sort_order INTEGER NOT NULL DEFAULT 0,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS customers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        phone TEXT NOT NULL UNIQUE,
        address TEXT,
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS orders (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_no TEXT UNIQUE,
        customer_id INTEGER NOT NULL,
        status TEXT NOT NULL DEFAULT 'مستلم',
        due_date TEXT,
        subtotal REAL NOT NULL DEFAULT 0,
        discount REAL NOT NULL DEFAULT 0,
        total REAL NOT NULL DEFAULT 0,
        paid REAL NOT NULL DEFAULT 0,
        payment_method TEXT NOT NULL DEFAULT 'نقدي',
        notes TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        delivered_at TEXT,
        FOREIGN KEY(customer_id) REFERENCES customers(id)
    );
    CREATE TABLE IF NOT EXISTS order_items (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        service_id INTEGER,
        service_name TEXT NOT NULL,
        quantity REAL NOT NULL DEFAULT 1,
        unit TEXT NOT NULL,
        unit_price REAL NOT NULL,
        line_total REAL NOT NULL,
        item_note TEXT,
        FOREIGN KEY(order_id) REFERENCES orders(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS expenses (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        expense_date TEXT NOT NULL,
        category TEXT NOT NULL,
        amount REAL NOT NULL,
        note TEXT,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER NOT NULL,
        amount REAL NOT NULL,
        payment_method TEXT NOT NULL DEFAULT 'نقدي',
        paid_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
        note TEXT,
        FOREIGN KEY(order_id) REFERENCES orders(id) ON DELETE CASCADE
    );
    CREATE TABLE IF NOT EXISTS cash_closings (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        closing_date TEXT NOT NULL UNIQUE,
        opening_cash REAL NOT NULL DEFAULT 0,
        actual_cash REAL,
        note TEXT,
        closed_at TEXT
    );
    CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        event_type TEXT NOT NULL,
        summary TEXT NOT NULL,
        created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS cash_days (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        business_id INTEGER NOT NULL,
        closing_date TEXT NOT NULL,
        opening_cash REAL NOT NULL DEFAULT 0,
        actual_cash REAL,
        total_received REAL NOT NULL DEFAULT 0,
        cash_received REAL NOT NULL DEFAULT 0,
        non_cash_received REAL NOT NULL DEFAULT 0,
        expense_total REAL NOT NULL DEFAULT 0,
        expected_cash REAL NOT NULL DEFAULT 0,
        cash_difference REAL NOT NULL DEFAULT 0,
        note TEXT,
        closed_at TEXT,
        UNIQUE(business_id,closing_date),
        FOREIGN KEY(business_id) REFERENCES businesses(id)
    );
    """
    with db() as connection:
        connection.executescript(schema)
        business_columns = {row[1] for row in connection.execute("PRAGMA table_info(businesses)")}
        if "subscription_status" not in business_columns:
            connection.execute("ALTER TABLE businesses ADD COLUMN subscription_status TEXT NOT NULL DEFAULT 'active'")
        if "subscription_end" not in business_columns:
            connection.execute("ALTER TABLE businesses ADD COLUMN subscription_end TEXT")
        if "last_seen" not in business_columns:
            connection.execute("ALTER TABLE businesses ADD COLUMN last_seen TEXT")
        if "setup_paid" not in business_columns:
            connection.execute("ALTER TABLE businesses ADD COLUMN setup_paid INTEGER NOT NULL DEFAULT 1")
        cash_day_columns = {row[1] for row in connection.execute("PRAGMA table_info(cash_days)")}
        for column in ("total_received", "cash_received", "non_cash_received", "expense_total", "expected_cash", "cash_difference"):
            if column not in cash_day_columns:
                connection.execute(f"ALTER TABLE cash_days ADD COLUMN {column} REAL NOT NULL DEFAULT 0")
        for table in ("services", "customers", "orders", "expenses", "payments", "audit_log"):
            columns = {row[1] for row in connection.execute(f"PRAGMA table_info({table})")}
            if "business_id" not in columns:
                connection.execute(f"ALTER TABLE {table} ADD COLUMN business_id INTEGER")
        customer_columns = {row[1] for row in connection.execute("PRAGMA table_info(customers)")}
        if "display_phone" not in customer_columns:
            connection.execute("ALTER TABLE customers ADD COLUMN display_phone TEXT")
        if "address" not in customer_columns:
            connection.execute("ALTER TABLE customers ADD COLUMN address TEXT")
        service_columns = {row[1] for row in connection.execute("PRAGMA table_info(services)")}
        if "treatment" not in service_columns:
            connection.execute("ALTER TABLE services ADD COLUMN treatment TEXT NOT NULL DEFAULT 'حسب الوصف'")
        connection.execute("UPDATE services SET treatment='غسيل وكوي' WHERE category='ملابس' AND (treatment IS NULL OR treatment='' OR treatment='حسب الوصف')")
        connection.execute("UPDATE services SET name='بنطال قماش' WHERE category='ملابس' AND name='بنطال'")
        connection.execute("UPDATE services SET name='فستان عادي' WHERE category='ملابس' AND name='فستان'")
        connection.execute("UPDATE services SET name='جاكيت / بليزر' WHERE category='ملابس' AND name='جاكيت'")
        connection.execute("UPDATE services SET category='سجاد وستائر' WHERE name='سجاد' AND category='مفروشات'")
        connection.execute("UPDATE services SET active=0 WHERE category='خدمات' AND name='كوي فقط'")
        for sort_order, service in enumerate(service_catalog_rows()):
            category, name, treatment, unit, price = service
            connection.execute(
                """INSERT INTO services(category,name,treatment,unit,price,sort_order)
                   SELECT ?,?,?,?,?,? WHERE NOT EXISTS (
                       SELECT 1 FROM services WHERE business_id IS NULL AND category=? AND name=? AND treatment=?
                   )""",
                (category, name, treatment, unit, price, sort_order, category, name, treatment),
            )
        business_ids = [row[0] for row in connection.execute("SELECT id FROM businesses")]
        for business_id in business_ids:
            for sort_order, service in enumerate(service_catalog_rows()):
                category, name, treatment, unit, price = service
                connection.execute(
                    """INSERT INTO services(business_id,category,name,treatment,unit,price,sort_order)
                       SELECT ?,?,?,?,?,?,? WHERE NOT EXISTS (
                           SELECT 1 FROM services WHERE business_id=? AND category=? AND name=? AND treatment=?
                       )""",
                    (business_id, category, name, treatment, unit, price, sort_order, business_id, category, name, treatment),
                )


def today():
    return date.today().isoformat()


def money(value):
    return f"{float(value or 0):,.2f}"


def normalized_phone(prefix, local_phone):
    digits = re.sub(r"\D", "", local_phone or "").lstrip("0")
    country = "972" if str(prefix).replace("+", "").replace("00", "") == "972" else "970"
    return f"+{country}{digits}", digits


app.jinja_env.filters["money"] = money


def current_business_id():
    return session.get("business_id")


def business_profile(connection):
    business_id = current_business_id()
    row = connection.execute("SELECT * FROM businesses WHERE id=?", (business_id,)).fetchone() if business_id else None
    if not row:
        return {"business_name": "FLEX", "business_phone": "", "business_address": "", "invoice_note": ""}
    return {"business_name": row["name"], "business_phone": row["phone"] or "", "business_address": row["address"] or "", "invoice_note": row["invoice_note"] or ""}


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.before_request
def enforce_business_subscription():
    if not session.get("user_id"):
        return None
    allowed = {"subscription", "logout", "static", "service_worker", "health"}
    if request.endpoint in allowed:
        return None
    with db() as connection:
        business = connection.execute(
            "SELECT subscription_status,subscription_end FROM businesses WHERE id=?",
            (current_business_id(),),
        ).fetchone()
        if not business:
            session.clear()
            return redirect(url_for("login"))
        expired = False
        if business["subscription_end"]:
            try:
                expired = datetime.fromisoformat(business["subscription_end"]) < datetime.now()
            except ValueError:
                expired = True
        if business["subscription_status"] != "active" or expired:
            if expired and business["subscription_status"] == "active":
                connection.execute(
                    "UPDATE businesses SET subscription_status='expired' WHERE id=?",
                    (current_business_id(),),
                )
            return redirect(url_for("subscription"))
    return None


@app.context_processor
def inject_business_profile():
    with db() as connection:
        return {"business": business_profile(connection)}


@app.get("/login")
def login():
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.post("/login")
def login_post():
    raw_phone = re.sub(r"[^0-9+]", "", request.form.get("phone", ""))
    phone, _ = normalized_phone(request.form.get("phone_prefix", "00970"), raw_phone)
    with db() as connection:
        user = connection.execute("SELECT * FROM users WHERE phone=? AND active=1", (phone,)).fetchone()
        if not user:
            user = connection.execute("SELECT * FROM users WHERE phone=? AND active=1", (raw_phone,)).fetchone()
    if not user or not check_password_hash(user["password_hash"], request.form.get("password", "")):
        flash("رقم الهاتف أو كلمة المرور غير صحيحة.", "error")
        return redirect(url_for("login"))
    session.clear()
    session.permanent = True
    session["user_id"] = user["id"]
    session["business_id"] = user["business_id"]
    session["role"] = user["role"]
    with db() as connection:
        connection.execute(
            "UPDATE businesses SET last_seen=? WHERE id=?",
            (datetime.now().isoformat(timespec="seconds"), user["business_id"]),
        )
        business = connection.execute(
            "SELECT subscription_status,subscription_end FROM businesses WHERE id=?",
            (user["business_id"],),
        ).fetchone()
    if not business or business["subscription_status"] != "active":
        return redirect(url_for("subscription"))
    return redirect(url_for("dashboard"))


@app.get("/register")
def register():
    return render_template("register.html")


@app.post("/register")
def register_post():
    business_name = request.form.get("business_name", "").strip()
    full_name = request.form.get("full_name", "").strip()
    phone, local_phone = normalized_phone(request.form.get("phone_prefix", "00970"), request.form.get("phone", ""))
    password = request.form.get("password", "")
    strong_password = len(password) >= 12 and re.search(r"[A-Z]", password) and re.search(r"[A-Za-z]", password) and re.search(r"\d", password)
    if not business_name or not full_name or len(local_phone) < 8 or not strong_password:
        flash("أكمل البيانات. كلمة المرور يجب أن تكون 12 خانة على الأقل وتضم حرفاً ورقماً وحرفاً إنجليزياً كبيراً.", "error")
        return redirect(url_for("register"))
    with db() as connection:
        if connection.execute("SELECT 1 FROM users WHERE phone=?", (phone,)).fetchone():
            flash("رقم الهاتف مسجل مسبقًا.", "error")
            return redirect(url_for("register"))
        business_id = connection.execute(
            "INSERT INTO businesses(name,phone,subscription_status,setup_paid) VALUES(?,?,'pending',0)",
            (business_name, phone),
        ).lastrowid
        user_id = connection.execute(
            "INSERT INTO users(business_id,full_name,phone,password_hash) VALUES(?,?,?,?)",
            (business_id, full_name, phone, generate_password_hash(password)),
        ).lastrowid
        templates = connection.execute("SELECT category,name,treatment,unit,price,sort_order FROM services WHERE business_id IS NULL").fetchall()
        connection.executemany(
            "INSERT INTO services(business_id,category,name,treatment,unit,price,sort_order) VALUES(?,?,?,?,?,?,?)",
            [(business_id, row["category"], row["name"], row["treatment"], row["unit"], row["price"], row["sort_order"]) for row in templates],
        )
    session.clear()
    session["user_id"] = user_id
    session["business_id"] = business_id
    session["role"] = "owner"
    flash("تم إنشاء الحساب. يصبح البرنامج جاهزًا بعد تأكيد الاشتراك من RT Studio.", "success")
    return redirect(url_for("subscription"))


@app.get("/subscription")
@login_required
def subscription():
    with db() as connection:
        business = connection.execute(
            "SELECT * FROM businesses WHERE id=?", (current_business_id(),)
        ).fetchone()
    return render_template("subscription.html", subscription=business)


@app.post("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.get("/")
@login_required
def dashboard():
    business_id = current_business_id()
    selected_date = request.args.get("date", today())
    query = request.args.get("q", "").strip()
    selected_customer_id = request.args.get("customer", type=int)
    with db() as connection:
        services = connection.execute(
            "SELECT * FROM services WHERE business_id=? AND active=1 ORDER BY category,sort_order,id", (business_id,)
        ).fetchall()
        selected_customer = connection.execute(
            "SELECT id,name,COALESCE(display_phone,'') phone,address FROM customers WHERE id=? AND business_id=?",
            (selected_customer_id, business_id),
        ).fetchone() if selected_customer_id else None
        params = [business_id]
        where = "WHERE o.business_id=?"
        if query:
            normalized = re.sub(r"[^0-9]", "", query).lstrip("0")
            where += " AND (o.order_no LIKE ? OR c.name LIKE ? OR c.display_phone LIKE ?)"
            params.extend([f"%{query}%", f"%{query}%", f"%{normalized or query}%"])
        orders = connection.execute(
            f"""SELECT o.*,c.name customer_name,COALESCE(c.display_phone,c.phone) customer_phone
                FROM orders o JOIN customers c ON c.id=o.customer_id
                {where} ORDER BY o.id DESC LIMIT 80""", params
        ).fetchall()
        order_summary = connection.execute(
            """SELECT COUNT(*) order_count,COALESCE(SUM(total),0) sales
               FROM orders WHERE business_id=? AND date(created_at)=?""", (business_id, selected_date)
        ).fetchone()
        received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=?", (business_id, selected_date)
        ).fetchone()[0]
        cash_received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=? AND payment_method='نقدي'",
            (business_id, selected_date),
        ).fetchone()[0]
        non_cash_received = received - cash_received
        outstanding = connection.execute(
            "SELECT COALESCE(SUM(total-paid),0) FROM orders WHERE business_id=? AND (status!='تم التسليم' OR total>paid)", (business_id,)
        ).fetchone()[0]
        summary = {"order_count": order_summary["order_count"], "sales": order_summary["sales"], "received": received, "outstanding": outstanding}
        expenses = connection.execute(
            "SELECT * FROM expenses WHERE business_id=? AND expense_date=? ORDER BY id DESC", (business_id, selected_date)
        ).fetchall()
        expense_total = sum(row["amount"] for row in expenses)
        closing = connection.execute(
            "SELECT * FROM cash_days WHERE business_id=? AND closing_date=?", (business_id, selected_date)
        ).fetchone()
        previous_closing = connection.execute(
            "SELECT actual_cash FROM cash_days WHERE business_id=? AND closing_date<? ORDER BY closing_date DESC LIMIT 1",
            (business_id, selected_date),
        ).fetchone()
        reminders = connection.execute(
            """SELECT o.id,o.order_no,o.due_date,o.status,c.name customer_name,c.phone customer_phone,
                      CAST((julianday(o.due_date)-julianday('now','localtime'))*24*60 AS INTEGER) minutes_left
               FROM orders o JOIN customers c ON c.id=o.customer_id
               WHERE o.business_id=? AND o.due_date IS NOT NULL AND o.status!='تم التسليم'
                 AND datetime(o.due_date) <= datetime('now','localtime','+3 hours')
               ORDER BY o.due_date LIMIT 20""", (business_id,)
        ).fetchall()
    grouped = {}
    for item in services:
        grouped.setdefault(item["category"], []).append(dict(item))
    suggested_opening_cash = previous_closing["actual_cash"] if previous_closing and previous_closing["actual_cash"] is not None else 0
    opening_cash = closing["opening_cash"] if closing else suggested_opening_cash
    cash_summary = {
        "total_received": received,
        "cash_received": cash_received,
        "non_cash_received": non_cash_received,
        "expense_total": expense_total,
        "opening_cash": opening_cash,
        "net_cash_movement": cash_received - expense_total,
        "expected_cash": opening_cash + cash_received - expense_total,
        "actual_cash": closing["actual_cash"] if closing else None,
        "difference": closing["cash_difference"] if closing else None,
    }
    return render_template(
        "dashboard.html", services=[dict(item) for item in services], grouped_services=grouped, orders=orders,
        summary=summary, expenses=expenses, expense_total=expense_total,
        selected_date=selected_date, query=query, closing=closing, cash_summary=cash_summary, reminders=reminders,
        selected_customer=selected_customer, auto_open_order=bool(selected_customer and request.args.get("new_order")),
    )


@app.get("/cash-accounts")
@login_required
def cash_accounts():
    business_id = current_business_id()
    selected_date = request.args.get("date", today())
    with db() as connection:
        order_summary = connection.execute(
            """SELECT COUNT(*) order_count,COALESCE(SUM(total),0) sales
               FROM orders WHERE business_id=? AND date(created_at)=?""",
            (business_id, selected_date),
        ).fetchone()
        total_received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=?",
            (business_id, selected_date),
        ).fetchone()[0]
        cash_received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=? AND payment_method='نقدي'",
            (business_id, selected_date),
        ).fetchone()[0]
        non_cash_received = total_received - cash_received
        expenses = connection.execute(
            "SELECT * FROM expenses WHERE business_id=? AND expense_date=? ORDER BY id DESC",
            (business_id, selected_date),
        ).fetchall()
        expense_total = sum(row["amount"] for row in expenses)
        closing = connection.execute(
            "SELECT * FROM cash_days WHERE business_id=? AND closing_date=?",
            (business_id, selected_date),
        ).fetchone()
        previous_closing = connection.execute(
            "SELECT actual_cash FROM cash_days WHERE business_id=? AND closing_date<? ORDER BY closing_date DESC LIMIT 1",
            (business_id, selected_date),
        ).fetchone()
    suggested_opening_cash = previous_closing["actual_cash"] if previous_closing and previous_closing["actual_cash"] is not None else 0
    opening_cash = closing["opening_cash"] if closing else suggested_opening_cash
    current_expected_cash = opening_cash + cash_received - expense_total
    closing_stale = bool(closing and (
        abs(float(closing["total_received"] or 0) - float(total_received or 0)) >= 0.01
        or abs(float(closing["cash_received"] or 0) - float(cash_received or 0)) >= 0.01
        or abs(float(closing["expense_total"] or 0) - float(expense_total or 0)) >= 0.01
    ))
    cash_summary = {
        "order_count": order_summary["order_count"],
        "sales": order_summary["sales"],
        "total_received": total_received,
        "cash_received": cash_received,
        "non_cash_received": non_cash_received,
        "expense_total": expense_total,
        "opening_cash": opening_cash,
        "net_cash_movement": cash_received - expense_total,
        "expected_cash": current_expected_cash,
        "current_difference": (float(closing["actual_cash"] or 0) - current_expected_cash) if closing else None,
        "closing_stale": closing_stale,
    }
    return render_template(
        "cash_accounts.html", selected_date=selected_date, cash_summary=cash_summary,
        expenses=expenses, closing=closing,
    )


@app.post("/orders")
@login_required
def create_order():
    business_id = current_business_id()
    name = request.form.get("customer_name", "").strip()
    phone = re.sub(r"[^0-9+]", "", request.form.get("customer_phone", ""))
    item_types = request.form.getlist("item_type[]")
    service_ids = request.form.getlist("service_id[]")
    manual_names = request.form.getlist("manual_name[]")
    item_units = request.form.getlist("item_unit[]")
    item_treatments = request.form.getlist("item_treatment[]")
    item_prices = request.form.getlist("item_price[]")
    quantities = request.form.getlist("quantity[]")
    item_notes = request.form.getlist("item_note[]")
    save_manual = request.form.getlist("save_manual[]")
    selected_customer_id = request.form.get("customer_id", type=int)
    if not selected_customer_id:
        flash("اختر زبوناً مسجلاً قبل إنشاء الطلب.", "error")
        return redirect(url_for("dashboard"))
    if not name or not phone or not item_types:
        flash("أدخل بيانات الزبون وأضف خدمة واحدة على الأقل.", "error")
        return redirect(url_for("dashboard"))
    with db() as connection:
        customer_key = f"{business_id}:{phone}"
        customer = connection.execute(
            "SELECT id,name,COALESCE(display_phone,'') phone FROM customers WHERE business_id=? AND id=?",
            (business_id, selected_customer_id),
        ).fetchone()
        if not customer:
            flash("ملف الزبون غير موجود.", "error")
            return redirect(url_for("dashboard"))
        customer_id = customer["id"]
        name, phone = customer["name"], customer["phone"]
        discount = max(float(request.form.get("discount") or 0), 0)
        paid = max(float(request.form.get("paid") or 0), 0)
        order_id = connection.execute(
            """INSERT INTO orders(business_id,customer_id,due_date,discount,paid,payment_method,notes)
               VALUES(?,?,?,?,?,?,?)""",
            (business_id, customer_id, request.form.get("due_date") or None, discount, paid,
             request.form.get("payment_method", "نقدي"), request.form.get("notes", "").strip()),
        ).lastrowid
        subtotal = 0
        for index, item_type in enumerate(item_types):
            quantity = max(float(quantities[index] or 1), 0.01)
            if item_type == "manual":
                service_name = (manual_names[index] if index < len(manual_names) else "").strip()
                unit = (item_units[index] if index < len(item_units) else "قطعة").strip() or "قطعة"
                treatment = (item_treatments[index] if index < len(item_treatments) else "حسب الوصف").strip() or "حسب الوصف"
                unit_price = max(float(item_prices[index] or 0), 0)
                service_id = None
                if not service_name:
                    continue
                if index < len(save_manual) and save_manual[index] == "1":
                    existing = connection.execute(
                        "SELECT id FROM services WHERE business_id=? AND name=? AND treatment=? AND unit=?",
                        (business_id, service_name, treatment, unit),
                    ).fetchone()
                    if existing:
                        service_id = existing["id"]
                    else:
                        service_id = connection.execute(
                            "INSERT INTO services(business_id,category,name,treatment,unit,price) VALUES(?,?,?,?,?,?)",
                            (business_id, "خدمات خاصة", service_name, treatment, unit, unit_price),
                        ).lastrowid
            else:
                raw_service_id = service_ids[index] if index < len(service_ids) else ""
                service = connection.execute(
                    "SELECT * FROM services WHERE id=? AND business_id=? AND active=1",
                    (raw_service_id, business_id),
                ).fetchone()
                if not service:
                    continue
                service_id = service["id"]
                service_name = service["name"]
                treatment = service["treatment"]
                unit = service["unit"]
                unit_price = service["price"]
            if unit == "قطعة":
                quantity = float(max(int(round(quantity)), 1))
            stored_service_name = service_name if treatment == "حسب الوصف" else f"{service_name} — {treatment}"
            line_total = quantity * unit_price
            item_note = (item_notes[index] if index < len(item_notes) else "").strip()
            subtotal += line_total
            connection.execute(
                """INSERT INTO order_items(order_id,service_id,service_name,quantity,unit,unit_price,line_total,item_note)
                   VALUES(?,?,?,?,?,?,?,?)""",
                (order_id, service_id, stored_service_name, quantity, unit, unit_price, line_total, item_note),
            )
        total = max(subtotal - discount, 0)
        paid = min(paid, total)
        order_no = f"FLEX-{order_id:06d}"
        connection.execute(
            "UPDATE orders SET order_no=?,subtotal=?,total=?,paid=? WHERE id=?",
            (order_no, subtotal, total, paid, order_id),
        )
        if paid > 0:
            connection.execute(
                "INSERT INTO payments(business_id,order_id,amount,payment_method,note) VALUES(?,?,?,?, 'دفعة إنشاء الطلب')",
                (business_id, order_id, paid, request.form.get("payment_method", "نقدي")),
            )
        connection.execute(
            "INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'order_created',?)",
            (business_id, f"إنشاء الطلب {order_no} بقيمة {total:.2f}"),
        )
    flash(f"تم حفظ الطلب {order_no} بنجاح.", "success")
    return redirect(url_for("order_detail", order_id=order_id))


@app.get("/api/customers")
@login_required
def customer_lookup():
    business_id = current_business_id()
    query = request.args.get("q", "").strip()
    if len(query) < 2:
        return jsonify(customers=[])
    term = f"%{query}%"
    digits = re.sub(r"[^0-9]", "", query).lstrip("0")
    phone_term = f"%{digits}%" if digits else term
    with db() as connection:
        customers = connection.execute(
            """SELECT c.id,c.name,COALESCE(c.display_phone,'') phone,COALESCE(c.address,'') address,
                      COUNT(o.id) order_count,
                      SUM(CASE WHEN o.id IS NOT NULL AND o.status!='تم التسليم' THEN 1 ELSE 0 END) open_count,
                      MAX(o.created_at) last_order_at
               FROM customers c LEFT JOIN orders o ON o.customer_id=c.id AND o.business_id=c.business_id
               WHERE c.business_id=? AND (c.name LIKE ? OR COALESCE(c.display_phone,'') LIKE ?)
               GROUP BY c.id,c.name,c.display_phone
               ORDER BY open_count DESC,last_order_at DESC LIMIT 10""",
            (business_id, term, phone_term),
        ).fetchall()
        result = []
        for customer in customers:
            open_orders = connection.execute(
                """SELECT id,order_no,status,due_date,total,paid FROM orders
                   WHERE business_id=? AND customer_id=? AND status!='تم التسليم'
                   ORDER BY id DESC LIMIT 5""",
                (business_id, customer["id"]),
            ).fetchall()
            result.append({
                "id": customer["id"], "name": customer["name"], "phone": customer["phone"], "address": customer["address"],
                "order_count": customer["order_count"], "open_count": customer["open_count"] or 0,
                "orders": [dict(order) for order in open_orders],
            })
    return jsonify(customers=result)


@app.get("/customers")
@login_required
def customers():
    with db() as connection:
        rows = connection.execute(
            """SELECT c.id,c.name,COALESCE(c.display_phone,'') phone,COALESCE(c.address,'') address,
                      COUNT(o.id) order_count,MAX(o.created_at) last_order_at
               FROM customers c LEFT JOIN orders o ON o.customer_id=c.id AND o.business_id=c.business_id
               WHERE c.business_id=? GROUP BY c.id ORDER BY c.name""",
            (current_business_id(),),
        ).fetchall()
    return render_template("customers.html", customers=rows)


@app.post("/customers")
@login_required
def create_customer():
    name = " ".join(request.form.get("name", "").split())
    prefix = request.form.get("phone_prefix", "00970")
    address = request.form.get("address", "").strip()
    phone, local_phone = normalized_phone(prefix, request.form.get("phone", ""))
    if len(name.split()) < 4 or len(local_phone) < 8 or not address:
        flash("أدخل الاسم الرباعي ورقم هاتف صحيح وعنوان السكن.", "error")
        return redirect(url_for("customers"))
    business_id = current_business_id()
    customer_key = f"{business_id}:{phone}"
    with db() as connection:
        if connection.execute("SELECT 1 FROM customers WHERE business_id=? AND phone=?", (business_id, customer_key)).fetchone():
            flash("رقم الهاتف مسجل لزبون سابق.", "error")
            return redirect(url_for("customers"))
        customer_id = connection.execute(
            "INSERT INTO customers(business_id,name,phone,display_phone,address) VALUES(?,?,?,?,?)",
            (business_id, name, customer_key, phone, address),
        ).lastrowid
    flash("تم إنشاء ملف الزبون. يمكنك الآن إضافة طلب له.", "success")
    return redirect(url_for("customer_detail", customer_id=customer_id))


@app.get("/customers/<int:customer_id>")
@login_required
def customer_detail(customer_id):
    with db() as connection:
        customer = connection.execute(
            "SELECT id,name,COALESCE(display_phone,'') phone,COALESCE(address,'') address FROM customers WHERE id=? AND business_id=?",
            (customer_id, current_business_id()),
        ).fetchone()
        if not customer:
            return "الزبون غير موجود", 404
        orders = connection.execute(
            "SELECT * FROM orders WHERE customer_id=? AND business_id=? ORDER BY id DESC",
            (customer_id, current_business_id()),
        ).fetchall()
    return render_template("customer.html", customer=customer, orders=orders)


@app.get("/orders/<int:order_id>")
@login_required
def order_detail(order_id):
    with db() as connection:
        order = connection.execute(
            """SELECT o.*,c.name customer_name,COALESCE(c.display_phone,c.phone) customer_phone
               FROM orders o JOIN customers c ON c.id=o.customer_id WHERE o.id=? AND o.business_id=?""", (order_id, current_business_id())
        ).fetchone()
        items = connection.execute("SELECT * FROM order_items WHERE order_id=?", (order_id,)).fetchall()
    if not order:
        return "الطلب غير موجود", 404
    return render_template("order.html", order=order, items=items)


@app.post("/orders/<int:order_id>/edit")
@login_required
def edit_order(order_id):
    business_id = current_business_id()
    names = request.form.getlist("item_name[]")
    quantities = request.form.getlist("quantity[]")
    units = request.form.getlist("unit[]")
    prices = request.form.getlist("unit_price[]")
    notes = request.form.getlist("item_note[]")
    prepared = []
    for index, raw_name in enumerate(names):
        name = raw_name.strip()
        if not name:
            continue
        unit = (units[index] if index < len(units) else "قطعة").strip() or "قطعة"
        try:
            quantity = max(float(quantities[index]), 0.01)
            unit_price = max(float(prices[index]), 0)
        except (ValueError, IndexError):
            flash("تحقق من الكمية والسعر في بنود الفاتورة.", "error")
            return redirect(url_for("order_detail", order_id=order_id))
        if unit == "قطعة":
            quantity = float(max(int(round(quantity)), 1))
        note = (notes[index] if index < len(notes) else "").strip()
        prepared.append((name, quantity, unit, unit_price, quantity * unit_price, note))
    if not prepared:
        flash("يجب أن تحتوي الفاتورة على بند واحد على الأقل.", "error")
        return redirect(url_for("order_detail", order_id=order_id))
    discount = max(float(request.form.get("discount") or 0), 0)
    subtotal = sum(item[4] for item in prepared)
    total = max(subtotal - discount, 0)
    try:
        due_date = requested_due_date()
    except (TypeError, ValueError):
        flash("تحقق من يوم وشهر وسنة ووقت التسليم.", "error")
        return redirect(url_for("order_detail", order_id=order_id))
    with db() as connection:
        order = connection.execute("SELECT * FROM orders WHERE id=? AND business_id=?", (order_id, business_id)).fetchone()
        if not order:
            return "الطلب غير موجود", 404
        if total < order["paid"]:
            flash("لا يمكن جعل إجمالي الفاتورة أقل من المبلغ المقبوض. عدّل الدفعات أولاً.", "error")
            return redirect(url_for("order_detail", order_id=order_id))
        connection.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
        connection.executemany(
            """INSERT INTO order_items(order_id,service_name,quantity,unit,unit_price,line_total,item_note)
               VALUES(?,?,?,?,?,?,?)""",
            [(order_id, name, quantity, unit, price, line_total, note) for name, quantity, unit, price, line_total, note in prepared],
        )
        connection.execute(
            "UPDATE orders SET due_date=?,discount=?,subtotal=?,total=?,notes=? WHERE id=?",
            (due_date, discount, subtotal, total, request.form.get("notes", "").strip(), order_id),
        )
        connection.execute(
            "INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'order_edited',?)",
            (business_id, f"تعديل الفاتورة {order['order_no']} من {order['total']:.2f} إلى {total:.2f}"),
        )
    flash("تم تحديث الفاتورة وإعادة احتساب المبلغ.", "success")
    return redirect(url_for("order_detail", order_id=order_id))


@app.post("/orders/<int:order_id>/due-date")
@login_required
def update_due_date(order_id):
    business_id = current_business_id()
    try:
        due_date = requested_due_date()
    except (TypeError, ValueError):
        flash("تحقق من يوم وشهر وسنة ووقت التسليم.", "error")
        return redirect(url_for("order_detail", order_id=order_id))
    with db() as connection:
        order = connection.execute(
            "SELECT order_no,due_date FROM orders WHERE id=? AND business_id=?",
            (order_id, business_id),
        ).fetchone()
        if not order:
            return "الطلب غير موجود", 404
        connection.execute(
            "UPDATE orders SET due_date=? WHERE id=? AND business_id=?",
            (due_date, order_id, business_id),
        )
        connection.execute(
            "INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'due_date_changed',?)",
            (business_id, f"تغيير موعد تسليم {order['order_no']} من {order['due_date'] or 'غير محدد'} إلى {due_date or 'غير محدد'}"),
        )
    flash("تم تحديث موعد التسليم.", "success")
    return redirect(url_for("order_detail", order_id=order_id))


@app.post("/orders/<int:order_id>/status")
@login_required
def update_status(order_id):
    business_id = current_business_id()
    status = request.form.get("status")
    if status not in {"مستلم", "قيد التنظيف", "جاهز", "تم التسليم", "إعادة تنظيف"}:
        return "حالة غير صحيحة", 400
    with db() as connection:
        delivered = datetime.now().isoformat(timespec="seconds") if status == "تم التسليم" else None
        changed = connection.execute("UPDATE orders SET status=?,delivered_at=COALESCE(?,delivered_at) WHERE id=? AND business_id=?", (status, delivered, order_id, business_id)).rowcount
        if not changed:
            return "الطلب غير موجود", 404
        connection.execute("INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'status_changed',?)", (business_id, f"تغيير حالة الطلب {order_id} إلى {status}"))
    flash("تم تحديث حالة الطلب.", "success")
    return redirect(url_for("order_detail", order_id=order_id))


@app.post("/orders/<int:order_id>/payments")
@login_required
def add_payment(order_id):
    business_id = current_business_id()
    amount = float(request.form.get("amount") or 0)
    with db() as connection:
        order = connection.execute("SELECT * FROM orders WHERE id=? AND business_id=?", (order_id, business_id)).fetchone()
        if not order:
            return "الطلب غير موجود", 404
        remaining = max(order["total"] - order["paid"], 0)
        if amount <= 0 or amount > remaining + 0.001:
            flash("قيمة الدفعة غير صحيحة أو أكبر من المبلغ المتبقي.", "error")
            return redirect(url_for("order_detail", order_id=order_id))
        method = request.form.get("payment_method", "نقدي")
        connection.execute("INSERT INTO payments(business_id,order_id,amount,payment_method,note) VALUES(?,?,?,?,?)", (business_id, order_id, amount, method, request.form.get("note", "").strip()))
        connection.execute("UPDATE orders SET paid=paid+? WHERE id=? AND business_id=?", (amount, order_id, business_id))
        connection.execute("INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'payment_added',?)", (business_id, f"قبض {amount:.2f} للطلب {order['order_no']}"))
    flash("تم تسجيل الدفعة في صندوق اليوم.", "success")
    return redirect(url_for("order_detail", order_id=order_id))


@app.post("/expenses")
@login_required
def create_expense():
    business_id = current_business_id()
    amount = float(request.form.get("amount") or 0)
    if amount <= 0:
        flash("أدخل مبلغ مصروف صحيحًا.", "error")
        return redirect(url_for("cash_accounts", date=request.form.get("expense_date", today())))
    with db() as connection:
        connection.execute(
            "INSERT INTO expenses(business_id,expense_date,category,amount,note) VALUES(?,?,?,?,?)",
            (business_id, request.form.get("expense_date", today()), request.form.get("category", "أخرى"), amount, request.form.get("note", "").strip()),
        )
    flash("تم تسجيل المصروف.", "success")
    return redirect(url_for("cash_accounts", date=request.form.get("expense_date", today())))


@app.post("/cash/close")
@login_required
def close_cash():
    business_id = current_business_id()
    selected_date = request.form.get("closing_date", today())
    actual = float(request.form.get("actual_cash") or 0)
    opening = float(request.form.get("opening_cash") or 0)
    with db() as connection:
        total_received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=?",
            (business_id, selected_date),
        ).fetchone()[0]
        cash_received = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM payments WHERE business_id=? AND date(paid_at)=? AND payment_method='نقدي'",
            (business_id, selected_date),
        ).fetchone()[0]
        expense_total = connection.execute(
            "SELECT COALESCE(SUM(amount),0) FROM expenses WHERE business_id=? AND expense_date=?",
            (business_id, selected_date),
        ).fetchone()[0]
        non_cash_received = total_received - cash_received
        expected_cash = opening + cash_received - expense_total
        cash_difference = actual - expected_cash
        connection.execute(
            """INSERT INTO cash_days(
                   business_id,closing_date,opening_cash,actual_cash,total_received,cash_received,
                   non_cash_received,expense_total,expected_cash,cash_difference,note,closed_at
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?) ON CONFLICT(business_id,closing_date) DO UPDATE SET
               opening_cash=excluded.opening_cash,actual_cash=excluded.actual_cash,
               total_received=excluded.total_received,cash_received=excluded.cash_received,
               non_cash_received=excluded.non_cash_received,expense_total=excluded.expense_total,
               expected_cash=excluded.expected_cash,cash_difference=excluded.cash_difference,
               note=excluded.note,closed_at=excluded.closed_at""",
            (
                business_id, selected_date, opening, actual, total_received, cash_received,
                non_cash_received, expense_total, expected_cash, cash_difference,
                request.form.get("note", ""), datetime.now().isoformat(timespec="seconds"),
            ),
        )
    flash("تم إقفال الصندوق وحفظ المبلغ الفعلي.", "success")
    return redirect(url_for("cash_accounts", date=selected_date))


@app.get("/settings/services")
@login_required
def service_settings():
    with db() as connection:
        services = connection.execute("SELECT * FROM services WHERE business_id=? ORDER BY category,sort_order,id", (current_business_id(),)).fetchall()
    categories = list(dict.fromkeys(service["category"] for service in services))
    return render_template("services.html", services=services, categories=categories)


@app.get("/settings/business")
@login_required
def business_settings():
    with db() as connection:
        profile = business_profile(connection)
    return render_template("business.html", profile=profile)


@app.post("/settings/business")
@login_required
def save_business_settings():
    values = {
        "business_name": request.form.get("business_name", "").strip(),
        "business_phone": request.form.get("business_phone", "").strip(),
        "business_address": request.form.get("business_address", "").strip(),
        "invoice_note": request.form.get("invoice_note", "").strip(),
    }
    if not values["business_name"]:
        flash("أدخل اسم المغسلة.", "error")
        return redirect(url_for("business_settings"))
    with db() as connection:
        connection.execute(
            "UPDATE businesses SET name=?,phone=?,address=?,invoice_note=? WHERE id=?",
            (values["business_name"], values["business_phone"], values["business_address"], values["invoice_note"], current_business_id()),
        )
        connection.execute(
            "INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'business_updated',?)",
            (current_business_id(), f"تحديث بيانات المغسلة: {values['business_name']}"),
        )
    flash("تم حفظ بيانات المغسلة وستظهر على الفواتير.", "success")
    return redirect(url_for("business_settings"))


@app.post("/settings/services")
@login_required
def add_service():
    with db() as connection:
        connection.execute(
            "INSERT INTO services(business_id,category,name,treatment,unit,price) VALUES(?,?,?,?,?,?)",
            (current_business_id(), request.form["category"].strip(), request.form["name"].strip(), request.form.get("treatment", "حسب الوصف").strip(), request.form["unit"], float(request.form["price"])),
        )
    flash("تمت إضافة الخدمة.", "success")
    return redirect(url_for("service_settings"))


@app.post("/settings/services/<int:service_id>")
@login_required
def edit_service(service_id):
    with db() as connection:
        old = connection.execute("SELECT * FROM services WHERE id=? AND business_id=?", (service_id, current_business_id())).fetchone()
        if not old:
            return "الخدمة غير موجودة", 404
        connection.execute(
            "UPDATE services SET category=?,name=?,treatment=?,unit=?,price=?,active=? WHERE id=? AND business_id=?",
            (request.form["category"].strip(), request.form["name"].strip(), request.form.get("treatment", "حسب الوصف").strip(), request.form["unit"], float(request.form["price"]), 1 if request.form.get("active") else 0, service_id, current_business_id()),
        )
        connection.execute("INSERT INTO audit_log(business_id,event_type,summary) VALUES(?,'price_changed',?)", (current_business_id(), f"تعديل {old['name']} من {old['price']:.2f} إلى {float(request.form['price']):.2f}"))
    flash("تم تحديث الخدمة والسعر الجديد سيطبق على الطلبات الجديدة فقط.", "success")
    return redirect(url_for("service_settings"))


@app.get("/reports/export.xlsx")
@login_required
def export_report():
    business_id = current_business_id()
    start = request.args.get("start", today())
    end = request.args.get("end", start)
    with db() as connection:
        orders = connection.execute(
            """SELECT o.*,c.name customer_name,COALESCE(c.display_phone,c.phone) customer_phone
               FROM orders o JOIN customers c ON c.id=o.customer_id
               WHERE o.business_id=? AND date(o.created_at) BETWEEN ? AND ? ORDER BY o.id""", (business_id, start, end)
        ).fetchall()
        expenses = connection.execute(
            "SELECT * FROM expenses WHERE business_id=? AND expense_date BETWEEN ? AND ? ORDER BY expense_date,id", (business_id, start, end)
        ).fetchall()
        payments = connection.execute(
            """SELECT p.*,o.order_no,c.name customer_name FROM payments p
               JOIN orders o ON o.id=p.order_id JOIN customers c ON c.id=o.customer_id
               WHERE p.business_id=? AND date(p.paid_at) BETWEEN ? AND ? ORDER BY p.paid_at,p.id""", (business_id, start, end)
        ).fetchall()
    workbook = Workbook()
    orders_sheet = workbook.active
    orders_sheet.title = "الطلبات"
    order_headers = ["رقم الطلب", "التاريخ", "الزبون", "الهاتف", "الحالة", "الإجمالي", "المدفوع", "المتبقي", "موعد التسليم"]
    orders_sheet.append(order_headers)
    for row in orders:
        orders_sheet.append([row["order_no"], row["created_at"], row["customer_name"], row["customer_phone"], row["status"], row["total"], row["paid"], row["total"]-row["paid"], row["due_date"]])
    expense_sheet = workbook.create_sheet("المصروفات")
    expense_sheet.append(["التاريخ", "التصنيف", "المبلغ", "الملاحظة"])
    for row in expenses:
        expense_sheet.append([row["expense_date"], row["category"], row["amount"], row["note"]])
    payments_sheet = workbook.create_sheet("المقبوضات")
    payments_sheet.append(["التاريخ", "رقم الطلب", "الزبون", "المبلغ", "طريقة الدفع", "الملاحظة"])
    for row in payments:
        payments_sheet.append([row["paid_at"], row["order_no"], row["customer_name"], row["amount"], row["payment_method"], row["note"]])
    summary_sheet = workbook.create_sheet("الملخص")
    sales = sum(row["total"] for row in orders)
    received = sum(row["amount"] for row in payments)
    expense_total = sum(row["amount"] for row in expenses)
    summary_sheet.append(["الفترة", f"{start} — {end}"])
    summary_sheet.append(["إجمالي المبيعات", sales])
    summary_sheet.append(["المقبوض", received])
    summary_sheet.append(["المتبقي", sales-received])
    summary_sheet.append(["المصروفات", expense_total])
    summary_sheet.append(["صافي الحركة النقدية", received-expense_total])
    for sheet in workbook.worksheets:
        sheet.sheet_view.rightToLeft = True
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="269FC2")
            cell.alignment = Alignment(horizontal="center")
        for column in sheet.columns:
            width = min(max(len(str(cell.value or "")) for cell in column) + 3, 34)
            sheet.column_dimensions[column[0].column_letter].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=f"FLEX-{start}-{end}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/api/scan")
@login_required
def scan_order():
    code = request.json.get("code", "").strip().upper()
    match = re.search(r"(?:FLEX-)?(\d{1,9})", code)
    if not match:
        return jsonify(ok=False, message="لم أتعرف على رقم الطلب"), 400
    order_id = int(match.group(1))
    with db() as connection:
        order = connection.execute("SELECT id,order_no,status FROM orders WHERE id=? AND business_id=?", (order_id, current_business_id())).fetchone()
    if not order:
        return jsonify(ok=False, message="الطلب غير موجود"), 404
    return jsonify(ok=True, url=url_for("order_detail", order_id=order["id"]), order_no=order["order_no"])


@app.get("/health")
def health():
    return jsonify(ok=True, app="FLEX")


@app.get("/sw.js")
def service_worker():
    response = send_from_directory(app.static_folder, "sw.js")
    response.headers["Cache-Control"] = "no-cache"
    response.headers["Service-Worker-Allowed"] = "/flex/"
    return response


init_db()

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=int(os.getenv("PORT", "8520")), debug=True)
